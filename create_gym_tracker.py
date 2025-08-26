import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timedelta
import calendar

def create_gym_tracker():
    # Create a new workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create the main workout tracker sheet
    ws = wb.create_sheet("Workout Tracker")
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    date_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Create headers
    headers = [
        "Date", "Day", "Training Part", "Exercise", "Sets", "Reps", 
        "Weight (lbs)", "Weight (kg)", "Rest Time", "Notes"
    ]
    
    # Add headers to worksheet
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Set column widths
    column_widths = [12, 8, 15, 20, 8, 8, 12, 12, 10, 25]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width
    
    # Create a sample month view (current month)
    current_date = datetime.now()
    year = current_date.year
    month = current_date.month
    
    # Get calendar for the month
    cal = calendar.monthcalendar(year, month)
    
    # Add month title
    month_name = calendar.month_name[month]
    ws.merge_cells('A1:J1')
    month_cell = ws.cell(row=1, column=1, value=f"{month_name} {year} - Workout Tracker")
    month_cell.font = Font(bold=True, size=16, color="FFFFFF")
    month_cell.fill = header_fill
    month_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Move headers to row 2
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add sample data rows for the month
    row_num = 3
    for week in cal:
        for day in week:
            if day != 0:  # Skip empty days
                date_obj = datetime(year, month, day)
                day_name = date_obj.strftime("%A")
                
                # Add date and day
                ws.cell(row=row_num, column=1, value=date_obj.strftime("%m/%d/%Y"))
                ws.cell(row=row_num, column=2, value=day_name)
                
                # Style the date cells
                for col in [1, 2]:
                    cell = ws.cell(row=row_num, column=col)
                    cell.fill = date_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Add empty cells for workout data
                for col in range(3, 11):
                    cell = ws.cell(row=row_num, column=col)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                row_num += 1
    
    # Create a summary sheet
    summary_ws = wb.create_sheet("Summary")
    
    # Summary headers
    summary_headers = [
        "Month", "Total Workouts", "Most Trained Part", "Total Sets", 
        "Total Reps", "Average Weight (lbs)", "Progress Notes"
    ]
    
    for col, header in enumerate(summary_headers, 1):
        cell = summary_ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Set summary column widths
    summary_widths = [12, 15, 20, 12, 12, 18, 30]
    for col, width in enumerate(summary_widths, 1):
        summary_ws.column_dimensions[chr(64 + col)].width = width
    
    # Create a template sheet
    template_ws = wb.create_sheet("Template")
    
    # Template headers
    template_headers = [
        "Training Part", "Exercise", "Sets", "Reps", "Weight (lbs)", 
        "Weight (kg)", "Rest Time", "Notes"
    ]
    
    for col, header in enumerate(template_headers, 1):
        cell = template_ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Set template column widths
    template_widths = [15, 20, 8, 8, 12, 12, 10, 25]
    for col, width in enumerate(template_widths, 1):
        template_ws.column_dimensions[chr(64 + col)].width = width
    
    # Add some common exercises to template
    common_exercises = [
        ["Chest", "Bench Press", "3-4", "8-12", "", "", "2-3 min", ""],
        ["Chest", "Incline Dumbbell Press", "3-4", "8-12", "", "", "2-3 min", ""],
        ["Back", "Pull-ups", "3-4", "8-12", "", "", "2-3 min", ""],
        ["Back", "Barbell Rows", "3-4", "8-12", "", "", "2-3 min", ""],
        ["Legs", "Squats", "3-4", "8-12", "", "", "2-3 min", ""],
        ["Legs", "Deadlifts", "3-4", "6-8", "", "", "3-4 min", ""],
        ["Shoulders", "Overhead Press", "3-4", "8-12", "", "", "2-3 min", ""],
        ["Arms", "Bicep Curls", "3-4", "10-15", "", "", "1-2 min", ""],
        ["Arms", "Tricep Dips", "3-4", "10-15", "", "", "1-2 min", ""],
        ["Core", "Planks", "3", "30-60 sec", "", "", "1 min", ""]
    ]
    
    for row, exercise in enumerate(common_exercises, 2):
        for col, value in enumerate(exercise, 1):
            cell = template_ws.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Save the workbook
    wb.save("gym_workout_tracker.xlsx")
    print("Gym workout tracker Excel file created successfully!")
    print("File saved as: gym_workout_tracker.xlsx")

if __name__ == "__main__":
    create_gym_tracker()
